import os
import pandas as pd
from tqdm import tqdm
import warnings
from openpyxl import load_workbook

warnings.simplefilter("ignore")  # Ignorar warnings de openpyxl

base_path = r"C:\Users\HeyCe\World Tree Technologies Inc\Forest Inventory - Documentos\USA\2024_ForestInventoryQ1_25\WT Cruises"
output_file = os.path.join(base_path, "Combined_Inventory.xlsx")
df_list = []

def read_input_sheet(file_path):
    """
    Intenta leer la hoja 'Input' (sin importar may├║sculas/min├║sculas) usando pd.read_excel.
    Si falla por estar protegida, utiliza openpyxl para cargar los datos, desactivando la protecci├│n.
    """
    try:
        # Intento normal usando pandas
        sheets = pd.read_excel(file_path, sheet_name=None)
        matched_sheet = next((sheets[s] for s in sheets if s.lower() == "input"), None)
        if matched_sheet is not None:
            return matched_sheet
    except Exception as e:
        # Si el mensaje de error menciona "protected", se intenta la lectura con openpyxl
        if "protected" in str(e).lower():
            try:
                wb = load_workbook(file_path, data_only=True)
                # Buscar la hoja cuyo nombre sea "Input" (sin distinci├│n entre may├║sculas/min├║sculas)
                target_sheet_name = next((s for s in wb.sheetnames if s.lower() == "input"), None)
                if target_sheet_name is not None:
                    ws = wb[target_sheet_name]
                    # Intentamos desactivar la protecci├│n. Algunas versiones permiten el m├®todo disable()
                    try:
                        ws.protection.disable()
                    except AttributeError:
                        ws.protection.sheet = False
                    # Convertir los valores de la hoja a una lista y luego a DataFrame
                    data = list(ws.values)
                    if not data:
                        return None
                    header = data[0]
                    df = pd.DataFrame(data[1:], columns=header)
                    return df
            except Exception as e2:
                print(f"[ERROR] Fall├│ la lectura alternativa para: {file_path}\n{e2}")
                return None
        else:
            print(f"[ERROR] No se pudo leer: {file_path}\n{e}")
            return None

# Recorrer los archivos Excel en el directorio
for root, dirs, files in os.walk(base_path):
    for file in files:
        if file.lower().endswith(".xlsx"):
            file_path = os.path.join(root, file)
            df = read_input_sheet(file_path)
            if df is not None:
                # Obtener el nombre del folder directo que contiene el archivo
                folder_name = os.path.basename(os.path.dirname(file_path))
                if " - " in folder_name:
                    farmer_name, contract_code = folder_name.split(" - ", 1)
                else:
                    farmer_name, contract_code = folder_name, "UNKNOWN"
                df.insert(0, "ContractCode", contract_code.strip())
                df.insert(0, "FarmerName", farmer_name.strip())
                df_list.append(df)

if df_list:
    final_df = pd.concat(df_list, ignore_index=True)
    final_df.to_excel(output_file, index=False)
    print(f"[OK] Archivo combinado guardado en:\n{output_file}")
else:
    print("[AVISO] No se encontraron hojas 'Input' o 'input' en los archivos.")
