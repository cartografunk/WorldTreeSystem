# MasterDatabaseManagement/Changes/new_contract_input_activation.py

from core.libs import pd, Path
from sqlalchemy import text
from core.db import get_engine
from core.paths import DATABASE_EXPORTS_DIR
from core.region import get_prefix
# Asegúrate que este módulo exponga extract_group_params y read_cell_by_key
from core.schema_helpers_db_management import extract_group_params, read_cell_by_key
from openpyxl import load_workbook

# === Config ===
CATALOG_FILE = Path(DATABASE_EXPORTS_DIR) / "changelog.xlsx"
# Si tu hoja real se llama "NewContractInputLog", cámbiala aquí:
SHEET_NAME = "NewContractInputLog_test"

engine = get_engine()


def _fetch_max(engine, prefix: str) -> int:
    """Máximo consecutivo actual por prefijo (US/MX/CR/GT)."""
    q = """
    SELECT MAX(CAST(SUBSTRING(contract_code,3) AS INT)) AS maxnum
    FROM masterdatabase.contract_farmer_information
    WHERE contract_code LIKE :pfx
    """
    df = pd.read_sql(q, engine, params={"pfx": f"{prefix}%"})
    return int(df.iloc[0]["maxnum"]) if not df.empty and pd.notna(df.iloc[0]["maxnum"]) else 0


# --- Parsers mínimos --------------------------------------------------------
def _to_int(v):
    n = pd.to_numeric(v, errors="coerce")
    return int(n) if pd.notna(n) else None

def _to_float(v):
    n = pd.to_numeric(v, errors="coerce")
    return float(n) if pd.notna(n) else None

def _to_date(v):
    if v is None or str(v).strip() == "":
        return None
    t = pd.to_datetime(v, errors="coerce", dayfirst=True)
    return t.date() if pd.notna(t) else None


# --- Helpers de validación (deben ir antes de main) -------------------------
REQUIRED_KEYS = ["region", "contractname", "plantingyear", "treescontract"]

def _reason_for_skip(row, headers, hdr_df):
    """Devuelve (razón, vals) si debe saltarse, o (None, vals) si está OK."""
    def _blank(x):
        return x is None or (isinstance(x, str) and x.strip() == "")

    missing = []
    vals = {}
    for k in REQUIRED_KEYS:
        v = read_cell_by_key(row, headers, hdr_df, k)
        vals[k] = v
        if _blank(v):
            missing.append(k)
    if missing:
        return f"faltan campos {missing}", vals

    # Prefijo válido
    pfx = get_prefix(vals["region"])
    if not pfx:
        return f"prefijo inválido para region='{vals['region']}'", vals

    # Números válidos
    tc = pd.to_numeric(vals["treescontract"], errors="coerce")
    if pd.isna(tc):
        return f"treescontract no numérico: {vals['treescontract']}", vals

    py = pd.to_numeric(vals["plantingyear"], errors="coerce")
    if pd.isna(py):
        return f"plantingyear no numérico: {vals['plantingyear']}", vals

    return None, vals  # OK


def main(dry_run: bool = False):
    wb = load_workbook(CATALOG_FILE)
    ws = wb[SHEET_NAME]

    # Headers del sheet (mantener orden para indexar celdas)
    headers = [c.value for c in ws[1]]
    hdr_df = pd.DataFrame(columns=headers)  # DF “de headers” para que los helpers resuelvan aliases del schema

    # --- Columnas de control (las ÚNICAS que escribimos) --------------------
    # Contract Code
    try:
        cc_idx = headers.index("Contract Code") + 1
    except ValueError:
        ws.cell(row=1, column=len(headers) + 1, value="Contract Code")
        headers.append("Contract Code")
        cc_idx = len(headers)

    # Done  (status tipo changelog: string “Done”, case-insensitive para leer; marcamos exactamente “Done”)
    try:
        done_idx = headers.index("Done") + 1
    except ValueError:
        ws.cell(row=1, column=len(headers) + 1, value="Done")
        headers.append("Done")
        done_idx = len(headers)

    def is_done(cell_value) -> bool:
        return bool(cell_value) and str(cell_value).strip().lower() == "done"

    # --- Transforms por grupo (aplicados por extract_group_params) ----------
    CFI_XFORM = {
        "phone": lambda x: str(x).strip() if x else None,
    }
    CTI_XFORM = {
        "plantingyear": _to_int,
        "harvest_year_10": _to_int,
        "treescontract": _to_int,
        "planted": _to_int,
        "plantingdate": _to_date,
        "latitude": _to_float,
        "longitude": _to_float,
    }

    # --- SQL (no destructivo) ----------------------------------------------
    sql_cfi = text("""
        INSERT INTO masterdatabase.contract_farmer_information
        (contract_code, contract_name, representative, farmer_number, phone, email, address,
         shipping_address, region, status, notes)
        VALUES
        (:contract_code, :contract_name, :representative, :farmer_number, :phone, :email, :address,
         :shipping_address, :region, :status, :notes)
        ON CONFLICT (contract_code) DO NOTHING
    """)

    sql_cti = text("""
        INSERT INTO masterdatabase.contract_tree_information
        (contract_code, planting_year, etp_year, harvest_year_10, trees_contract, planted, strain,
         planting_date, species, latitude, longitude, land_location_gps)
        VALUES
        (:contract_code, :planting_year, :planting_year, :harvest_year_10, :trees_contract, :planted, :strain,
         :planting_date, :species, :latitude, :longitude, :land_location_gps)
        ON CONFLICT (contract_code) DO NOTHING
    """)

    # --- Estado de corrida --------------------------------------------------
    counters = {}  # prefijo -> último consecutivo asignado localmente
    applied = 0
    failed = 0

    # Info rápida de la hoja
    df = pd.DataFrame(ws.values)
    print(f"📄 Hoja {SHEET_NAME} tiene {len(df) - 1} filas de datos y {len(df.columns)} columnas")

    # Acumuladores de preview en dry-run
    to_cfi_preview, to_cti_preview = [], []

    # --- Loop de filas: respeta orden del sheet -----------------------------
    for r, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        print(f"➡️  Fila {r}: inspeccionando…")

        # Skip si ya está Done (igual que changelog_activation)
        done_cell = ws.cell(row=r, column=done_idx)
        if is_done(done_cell.value):
            print(f"⏭️  Fila {r} saltada: ya estaba 'Done'")
            continue

        # Requisitos mínimos y validaciones
        reason, vals = _reason_for_skip(row, headers, hdr_df)
        if reason:
            print(f"⛔ Fila {r} descartada: {reason} | vals={vals}")
            continue

        region         = vals["region"]
        contract_name  = vals["contractname"]
        planting_year  = vals["plantingyear"]
        trees_contract = vals["treescontract"]

        # Contract Code: respeta si ya existe; si no, serializa por prefijo en orden de fila
        cc_cell = ws.cell(row=r, column=cc_idx)
        contract_code = (cc_cell.value or "").strip()
        if not contract_code:
            pfx = get_prefix(region)
            if pfx not in counters:
                counters[pfx] = _fetch_max(engine, pfx)
            counters[pfx] += 1
            contract_code = f"{pfx}{counters[pfx]:04d}"
            print(f"🆕 Fila {r}: asignado contract_code={contract_code}")
            if not dry_run:
                cc_cell.value = contract_code  # solo esta celda

        # Construcción de parámetros usando helpers por grupo
        cfi_params = extract_group_params(row, headers, hdr_df, "cfi", transforms=CFI_XFORM)
        cti_params = extract_group_params(row, headers, hdr_df, "cti", transforms=CTI_XFORM)

        # Ajustes CTI: planting_year/etp_year/harvest_year_10
        py_num = _to_int(planting_year)
        cti_params["planting_year"] = py_num
        if cti_params.get("harvest_year_10") is None and py_num is not None:
            cti_params["harvest_year_10"] = py_num + 10

        print(f"🧩 Fila {r}: ready | region={region} | name={contract_name} | py={py_num} | trees={trees_contract}")

        # Inserciones o preview
        if dry_run:
            to_cfi_preview.append({"contract_code": contract_code, **cfi_params})
            to_cti_preview.append({"contract_code": contract_code, **cti_params})
        else:
            try:
                with engine.begin() as conn:
                    conn.execute(sql_cfi, {"contract_code": contract_code, **cfi_params})
                    conn.execute(sql_cti, {"contract_code": contract_code, **cti_params})
                done_cell.value = "Done"
                applied += 1
                print(f"✅ Fila {r} aplicada y marcada Done")
            except Exception as e:
                failed += 1
                print(f"💥 Fila {r} error: {e}")

    # Guardado si no es dry-run
    if not dry_run:
        wb.save(CATALOG_FILE)

    # Preview si es dry-run
    if dry_run:
        if to_cfi_preview or to_cti_preview:
            print("=== Preview CFI (top 3) ===")
            print(pd.DataFrame(to_cfi_preview).head(3))
            print("=== Preview CTI (top 3) ===")
            print(pd.DataFrame(to_cti_preview).head(3))
        else:
            print("🪵 No hay preview: todas las filas fueron descartadas (revisa razones arriba).")

    print("✅ Contract Code serializado por prefijo y orden de fila.")
    print(f"✅ Filas marcadas como 'Done': {applied} | ❌ fallidos: {failed} | dry_run={dry_run}")


if __name__ == "__main__":
    main(dry_run=False)
