# core/sheets.py
from __future__ import annotations
from typing import Dict, List, Optional, Iterable, Tuple
from openpyxl import load_workbook
from core.libs import pd
from core.schema_helpers_db_management import read_cell_by_key
from core.schema_helpers import clean_column_name
from core.libs import Path

# =========================
#   ESTATUS CENTRALIZADO
# =========================
STATUS_READY = "Ready"
STATUS_DONE  = "Done"

def _norm_text(v) -> str:
    """Normaliza texto de celda: quita NBSP, trim y lower."""
    s = "" if v is None else str(v)
    s = s.replace("\u00A0", " ")  # NBSP -> espacio normal
    return s.strip()

def is_ready(v) -> bool:
    return _norm_text(v).lower() == STATUS_READY.lower()

def is_done(v) -> bool:
    return _norm_text(v).lower() == STATUS_DONE.lower()

# =========================
#   HOJAS (Excel helper)
# =========================
class Sheet:
    """
    Utilidad central para trabajar con una hoja Excel (openpyxl) usando schema:
    - Carga libro/hoja
    - Mantiene headers + hdr_df (para get_column/read_cell_by_key)
    - ensure_column(): crea columna si falta y devuelve índice (1-based)
    - ensure_status_column(): idem pero para columna de control (p.ej. 'change_in_db')
    - read(): lee por key lógica usando schema/aliases
    - index_of(): busca columna por nombre o nombre normalizado
    - iter_rows(): itera filas de datos (2..max_row)
    - iter_ready_rows(status_col_idx): itera solo filas con status == Ready
    - mark_status()/mark_done(): setea el status en la celda
    - save(): guarda el archivo
    """
    def __init__(self, path, sheet_name: str):
        self.path = Path(path)
        self.wb = load_workbook(self.path)
        if sheet_name not in self.wb.sheetnames:
            raise ValueError(f"❌ La hoja '{sheet_name}' no existe en {self.path}")
        self.ws = self.wb[sheet_name]
        self.headers: List[str] = [c.value if c.value is not None else "" for c in self.ws[1]]
        self.hdr_df = pd.DataFrame(columns=self.headers)
        self._rebuild_header_map()

    # ---------- internals ----------
    def _rebuild_header_map(self):
        self.header_map: Dict[str, int] = {h: i + 1 for i, h in enumerate(self.headers)}
        self.header_map_norm: Dict[str, int] = {clean_column_name(h): i for i, h in enumerate(self.headers)}

    # ---------- columnas ----------
    def ensure_column(self, name: str) -> int:
        """Si no existe, crea la columna al final. Devuelve índice 1-based."""
        if name in self.header_map:
            return self.header_map[name]
        col = len(self.headers) + 1
        self.ws.cell(row=1, column=col, value=name)
        self.headers.append(name)
        self._rebuild_header_map()
        return col

    def ensure_status_column(self, name: str = "change_in_db") -> int:
        """Crea/obtiene la columna de estatus de control."""
        return self.ensure_column(name)

    def index_of(self, header_name: str) -> Optional[int]:
        """Índice 1-based por nombre exacto o normalizado (clean_column_name)."""
        if header_name in self.header_map:
            return self.header_map[header_name]
        norm = clean_column_name(header_name)
        idx0 = self.header_map_norm.get(norm)
        return (idx0 + 1) if idx0 is not None else None

    # ---------- lectura de filas ----------
    def iter_rows(self):
        """Itera filas de datos: (row_index, openpyxl_row)."""
        for r, row in enumerate(self.ws.iter_rows(min_row=2, max_row=self.ws.max_row), start=2):
            yield r, row

    def iter_ready_rows(self, status_col_idx: int):
        """Itera solo filas con status == STATUS_READY en la columna dada."""
        for r, row in self.iter_rows():
            v = self.get_cell(r, status_col_idx).value
            if is_ready(v):
                yield r, row

    # ---------- lectura por schema ----------
    def read(self, row, logical_key: str):
        """Lee valor por key lógica usando schema/aliases (read_cell_by_key)."""
        try:
            return read_cell_by_key(row, self.headers, self.hdr_df, logical_key)
        except Exception:
            return None

    # ---------- celdas ----------
    def get_cell(self, r: int, c: int):
        return self.ws.cell(row=r, column=c)

    def mark_status(self, r: int, status_col_idx: int, value: str):
        self.ws.cell(row=r, column=status_col_idx, value=value)

    def mark_done(self, r: int, status_col_idx: int):
        """Alias útil; marca Done en la columna de status."""
        self.mark_status(r, status_col_idx, STATUS_DONE)

    # ---------- persistencia ----------
    def save(self):
        self.wb.save(self.path)

# =========================
#   Catálogos ChangeLog
# =========================
def read_changelog_catalogs(catalog_file) -> tuple[pd.DataFrame, pd.DataFrame]:
    fields = pd.read_excel(catalog_file, sheet_name="FieldsCatalog")
    reasons = pd.read_excel(catalog_file, sheet_name="ChangeReasonsCatalog")
    return fields, reasons

def get_table_for_field(fields_catalog: pd.DataFrame, field: str) -> str | None:
    matches = fields_catalog[fields_catalog["target_field"] == field]
    return matches["target_table"].iloc[0] if not matches.empty else None

# =========================
#   DF utils / export
# =========================
def remove_tz(df: pd.DataFrame) -> pd.DataFrame:
    for col in df.columns:
        if pd.api.types.is_datetime64tz_dtype(df[col]):
            df[col] = df[col].dt.tz_localize(None)
    return df

def export_tables_to_excel(engine, tables: Iterable[str], out_path, order_by: str = "contract_code"):
    """
    Exporta tablas SQL a un Excel (una hoja por tabla). Sobrescribe el archivo.
    """
    out_path = Path(out_path)
    print("⏳ Sobrescribiendo Excel...")
    with pd.ExcelWriter(out_path, engine="openpyxl", mode="w") as writer:
        for table in tables:
            df = pd.read_sql(f'SELECT * FROM masterdatabase."{table}"', engine)
            df = remove_tz(df)
            if order_by in df.columns:
                df = df.sort_values(order_by)
            df.to_excel(writer, sheet_name=table[:31], index=False)
            print(f"Exportado: {table}")
    print(f"✅ Exportación finalizada: {out_path}")
