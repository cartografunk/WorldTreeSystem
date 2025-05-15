import os
import pandas as pd
from tqdm import tqdm
import warnings
from openpyxl import load_workbook

warnings.simplefilter("ignore")  # Ignorar warnings de openpyxl

# Lista de códigos a reprocesar (extraída de la información entregada)
reprocess_contract_codes = [
    "US0145",
    "US0150",
    "US0155",
    "US0156",
    "US0157",
    "US0161",
    "US0163",
    "US0164",
    "US0165"
]

base_path = r"C:\Users\HeyCe\World Tree Technologies Inc\Forest Inventory - Documentos\USA\2024_ForestInventoryQ1_25\WT Cruises"
output_file = os.path.join(base_path, "Combined_Inventory_Reprocess.xlsx")
df_list = []

def read_input_sheet(file_path):
    """
    Intenta leer la hoja 'Input' (ignorando mayúsculas/minúsculas) usando pandas.
    Si la hoja está protegida, usa openpyxl para desactivar la protección y extraer los datos.
    """
    try:
        sheets = pd.read_excel(file_path, sheet_name=None)
        matched_sheet = next((sheets[s] for s in sheets if s.lower() == "input"), None)
        if matched_sheet is not None:
            return matched_sheet
    except Exception as e:
        # Si el error menciona 'protected', se intenta con openpyxl
        if "protected" in str(e).lower():
            try:
                wb = load_workbook(file_path, data_only=True)
                target_sheet_name = next((s for s in wb.sheetnames if s.lower() == "input"), None)
                if target_sheet_name is not None:
                    ws = wb[target_sheet_name]
                    # Desactivar la protección según la versión de openpyxl
                    try:
                        ws.protection.disable()
                    except AttributeError:
                        ws.protection.sheet = False
                    # Convertir los datos de la hoja a lista y luego a DataFrame
                    data = list(ws.values)
                    if not data:
                        return None
                    header = data[0]
                    df = pd.DataFrame(data[1:], columns=header)
                    return df
            except Exception as e2:
                print(f"[ERROR] Falló la lectura alternativa para: {file_path}\n{e2}")
                return None
        else:
            print(f"[ERROR] No se pudo leer: {file_path}\n{e}")
            return None

# Recorrer los archivos Excel dentro del directorio base
for root, dirs, files in os.walk(base_path):
    for file in files:
        if file.lower().endswith(".xlsx"):
            file_path = os.path.join(root, file)
            # Obtener el nombre del folder contenedor
            folder_name = os.path.basename(os.path.dirname(file_path))
            # Intentar dividir usando " - " o, si no, usar "-" como separador
            if " - " in folder_name:
                parts = folder_name.split(" - ")
            elif "-" in folder_name:
                parts = folder_name.split("-")
            else:
                parts = [folder_name]

            if len(parts) >= 2:
                farmer_name = parts[0].strip()
                contract_code = parts[1].strip()
            else:
                farmer_name = folder_name.strip()
                contract_code = "UNKNOWN"

            # Filtrar solo los archivos cuyo contract code esté en la lista de reprocesamiento
            if contract_code not in reprocess_contract_codes:
                continue

            # Leer la hoja "Input"
            df = read_input_sheet(file_path)
            if df is not None:
                # Insertar columnas adicionales según el folder
                df.insert(0, "ContractCode", contract_code)
                df.insert(0, "FarmerName", farmer_name)
                df_list.append(df)
            else:
                print(f"[AVISO] No se pudo extraer la hoja 'Input' en: {file_path}")

if df_list:
    final_df = pd.concat(df_list, ignore_index=True)
    final_df.to_excel(output_file, index=False)
    print(f"[OK] Archivo combinado de reproceso guardado en:\n{output_file}")
else:
    print("[AVISO] No se encontraron hojas 'Input' en los archivos a reprocesar.")
