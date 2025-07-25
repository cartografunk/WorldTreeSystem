from core.libs import pd, openpyxl, shutil, text
from core.db import get_engine
from openpyxl import load_workbook

# Rutas de los archivos
EXCEL_FILE = r"C:\Users\HeyCe\World Tree Technologies Inc\Operations - Documentos\WorldTreeSystem\DatabaseExports\masterdatabase_export.xlsx"
CATALOG_FILE = r"C:\Users\HeyCe\World Tree Technologies Inc\Operations - Documentos\WorldTreeSystem\DatabaseExports\changelog.xlsx"
EXCEL_BACKUP = EXCEL_FILE.replace(".xlsx", "_backup.xlsx")
SHEET_NAME = "ChangeLog"

engine = get_engine()

def backup_excel():
    shutil.copyfile(EXCEL_FILE, EXCEL_BACKUP)
    print(f"📝 Backup creado: {EXCEL_BACKUP}")

def read_catalogs():
    fields = pd.read_excel(CATALOG_FILE, sheet_name="FieldsCatalog")
    reasons = pd.read_excel(CATALOG_FILE, sheet_name="ChangeReasonsCatalog")
    return fields, reasons

def get_table_for_field(fields_catalog, field):
    matches = fields_catalog[fields_catalog['target_field'] == field]
    if not matches.empty:
        return matches['target_table'].iloc[0]
    return None

def process_changelog_and_update_sql(fields_catalog):
    wb = load_workbook(CATALOG_FILE)
    ws = wb[SHEET_NAME]
    # Detecta encabezados
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {col: idx+1 for idx, col in enumerate(header_row)}
    code_col = header_map.get("contract_code")
    field_col = header_map.get("target_field")
    change_col = header_map.get("change")
    status_col = header_map.get("change_in_db")
    changes_applied = 0

    # Procesa solo filas donde status_col está vacío
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        done = row[status_col-1].value
        if done and str(done).strip().lower() == "done":
            continue
        contract_code = row[code_col-1].value
        target_field = row[field_col-1].value
        change = row[change_col-1].value
        if not contract_code or not target_field:
            continue
        table = get_table_for_field(fields_catalog, target_field)
        if not table:
            print(f"❌ Campo '{target_field}' no encontrado en FieldsCatalog")
            continue
        # Detección de tipo de dato (numérico o string)
        if isinstance(change, (int, float)) or (isinstance(change, str) and change.replace('.', '', 1).isdigit()):
            change_value = change
        else:
            change_value = f"'{change}'"
        sql = f"""
            UPDATE masterdatabase.{table}
            SET {target_field} = {change_value}
            WHERE contract_code = '{contract_code}'
        """
        print("Ejecutando:", sql)
        with engine.begin() as conn:
            conn.execute(text(sql))
        # Marca la celda como Done
        ws.cell(row=i, column=status_col, value="Done")
        changes_applied += 1
    wb.save(CATALOG_FILE)
    print(f"✅ Cambios aplicados: {changes_applied}. Solo columna 'change_in_db' modificada (dropdowns conservados).")

def remove_tz(df):
    for col in df.select_dtypes(include=["datetimetz"]).columns:
        df[col] = df[col].dt.tz_localize(None)
    return df

def refresh_tables_in_excel(ordered_tables):
    """Sobrescribe las hojas, solo las tablas de datos."""
    print("⏳ Sobrescribiendo Excel...")
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="w") as writer:
        for table in ordered_tables:
            df = pd.read_sql(f'SELECT * FROM masterdatabase."{table}"', engine)
            df = remove_tz(df)
            if "contract_code" in df.columns:
                df = df.sort_values("contract_code")
            df.to_excel(writer, sheet_name=table[:31], index=False)
            print(f"Exportado: {table}")
    print(f"✅ Exportación finalizada: {EXCEL_FILE}")

def main():
    ORDERED_TABLES = [
        "contract_tree_information",
        "contract_farmer_information",
        "contract_allocation",
        "inventory_metrics",
        "inventory_metrics_current"
    ]
    print("💾 Realizando backup del Excel original...")
    backup_excel()

    print("📚 Leyendo catálogos...")
    fields_catalog, _ = read_catalogs()

    print("🚩 Aplicando cambios pendientes...")
    process_changelog_and_update_sql(fields_catalog)

    print("💾 Re-escribiendo Excel actualizado...")
    refresh_tables_in_excel(ORDERED_TABLES)

    print("🏁 Flujo completo terminado.")

if __name__ == "__main__":
    main()
