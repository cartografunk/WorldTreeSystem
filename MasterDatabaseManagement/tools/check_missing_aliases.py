# MasterDatabaseManagement/tools/check_missing_aliases.py
# -*- coding: utf-8 -*-
"""
Chequea, usando get_column, si las keys lógicas resuelven contra los headers reales
de tu sheet. Si no, sugiere los aliases exactos que deberías agregar en core.schema.COLUMNS.

Uso:
  python -m MasterDatabaseManagement.tools.check_missing_aliases
  # Opcional:
  #   --sheet NewContractInputLog
  #   --file  "C:\\...\\DatabaseExports\\changelog.xlsx"
"""

from __future__ import annotations
import argparse
from typing import List, Dict

from openpyxl import load_workbook

from core.libs import pd, Path
from core.paths import DATABASE_EXPORTS_DIR
from core.schema_helpers import get_column  # <— ya existente

DEFAULT_XLSX = Path(DATABASE_EXPORTS_DIR) / "changelog.xlsx"
DEFAULT_SHEET = "NewContractInputLog_test"

# 👇 Cabe tal cual con los headers reales que vimos en tu extracto
EXPECTED_HEADERS_BY_KEY: Dict[str, List[str]] = {
    "region":        ["Region"],
    "contractname":  ["Contract Name"],
    "plantingyear":  ["Planting Year", "ETP Year"],
    "treescontract": ["#TreesContract", "Trees Contract", "TreesContract"],
}


def read_headers(xlsx_path: Path, sheet_name: str) -> List[str]:
    wb = load_workbook(xlsx_path, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"❌ La hoja '{sheet_name}' no existe en {xlsx_path}")
    ws = wb[sheet_name]
    headers = [c.value if c.value is not None else "" for c in ws[1]]
    return headers


def main():
    ap = argparse.ArgumentParser("check_missing_aliases")
    ap.add_argument("--file", type=str, default=str(DEFAULT_XLSX),
                    help=f"Ruta del XLSX (default: {DEFAULT_XLSX})")
    ap.add_argument("--sheet", type=str, default=DEFAULT_SHEET,
                    help=f"Hoja a inspeccionar (default: {DEFAULT_SHEET})")
    args = ap.parse_args()

    xlsx_path = Path(args.file)
    headers = read_headers(xlsx_path, args.sheet)
    hdr_df = pd.DataFrame(columns=headers)  # patrón que ya usas con get_column

    print(f"📄 Archivo: {xlsx_path}")
    print(f"🗂️  Hoja: {args.sheet}")
    print(f"🧾 Headers ({len(headers)}): {headers}\n")

    missing_any = False
    print("🔎 Checando keys → header (via get_column):")
    for key, candidates in EXPECTED_HEADERS_BY_KEY.items():
        try:
            real = get_column(key, hdr_df)  # intenta resolver con schema (key/sql_name/aliases)
            print(f"  ✅ {key:>13}  ←  {real}")
        except Exception:
            # No resolvió con los aliases actuales → veamos qué header existe y sugiere alias
            present = [h for h in candidates if h in headers]
            if present:
                missing_any = True
                # Sugerir agregar el primero que exista (o todos, si quieres)
                print(f"  ❌ {key:>13}  ←  (no resuelto)  ➜  agrega alias en COLUMNS: {present}")
            else:
                print(f"  ⚠️  {key:>13}  ←  (header esperado no está en la sheet)  candidatos: {candidates}")

    if missing_any:
        print("\n📝 Copia/pega en core/schema.py (COLUMNS) los aliases que faltan, ej.:")
        print("  {'key': 'region',        'sql_name': 'region',        'aliases': ['Region']}")
        print("  {'key': 'contractname',  'sql_name': 'contract_name', 'aliases': ['Contract Name']}")
        print("  {'key': 'plantingyear',  'sql_name': 'planting_year', 'aliases': ['Planting Year', 'ETP Year']}")
        print("  {'key': 'treescontract', 'sql_name': 'trees_contract','aliases': ['#TreesContract', 'Trees Contract', 'TreesContract']}")
    else:
        print("\n🎉 Todo bien: tus aliases actuales en COLUMNS ya resuelven contra la sheet.")


if __name__ == "__main__":
    main()
