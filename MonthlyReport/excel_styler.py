#MonthlyReport/excel_styler.py

# -*- coding: utf-8 -*-
# MonthlyReport/excel_styler.py
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import date
from core.paths import MONTHLY_REPORT_DIR, ensure_all_paths_exist
import pandas as pd, shutil

def _find_header_anchor(ws, columns, max_scan_rows=100, max_scan_cols=100):
    cols, n = list(columns), len(columns)
    for r in range(1, max_scan_rows + 1):
        for c in range(1, max_scan_cols - n + 2):
            if all(ws.cell(row=r, column=c+j).value == cols[j] for j in range(n)):
                return r, c
    return None, None

def _inside(ref, row, col):
    from openpyxl.utils.cell import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    return (min_row <= row <= max_row) and (min_col <= col <= max_col)

def _resize_table_if_needed(ws, header_row, col_start, n_rows, n_cols):
    if not hasattr(ws, "tables"):
        return
    start = get_column_letter(col_start) + str(header_row)
    end   = get_column_letter(col_start + n_cols - 1) + str(header_row + n_rows)
    new_ref = f"{start}:{end}"
    for tbl in ws.tables.values():
        if _inside(tbl.ref, header_row, col_start):
            tbl.ref = new_ref

def _measure_current_block(ws, header_row, col_start, n_cols, max_rows=100000):
    r, count = header_row + 1, 0
    while count < max_rows:
        row_vals = [ws.cell(row=r+count, column=col_start+j).value for j in range(n_cols)]
        if all(v in (None, "") for v in row_vals):
            break
        count += 1
    return count

def _clear_old_block(ws, header_row, col_start, n_cols, old_n_rows):
    if old_n_rows <= 0: return
    for rr in range(header_row+1, header_row+old_n_rows+1):
        for j in range(n_cols):
            ws.cell(row=rr, column=col_start+j).value = None

def write_df_preserving_style(ws, df: pd.DataFrame):
    """
    Escribe un DataFrame en una hoja de Excel preservando los estilos existentes.
    Ahora maneja columnas que pueden estar en el template pero no en el DF y viceversa.
    """
    header_row = None
    header_row_idx = None

    # Buscar la fila de encabezados
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        row_values = [str(cell).strip() if cell is not None else '' for cell in row]
        if any(col in row_values for col in df.columns):
            header_row = row_values
            header_row_idx = row_idx
            break

    if header_row is None:
        raise ValueError(f"No encontré ningún encabezado del DF en hoja '{ws.title}'")

    # Mapear columnas del DF a posiciones en Excel
    col_mapping = {}
    for df_col in df.columns:
        if df_col in header_row:
            excel_col_idx = header_row.index(df_col)
            col_mapping[df_col] = excel_col_idx

    # Verificar que al menos algunas columnas coincidan
    if not col_mapping:
        raise ValueError(f"Ninguna columna del DF coincide con la plantilla en hoja '{ws.title}'")

    print(f"✅ Hoja '{ws.title}': {len(col_mapping)}/{len(df.columns)} columnas mapeadas")

    # Limpiar datos existentes (mantener solo header)
    ws.delete_rows(header_row_idx + 1, ws.max_row)

    # Escribir datos del DataFrame
    for df_row_idx, df_row in df.iterrows():
        excel_row_idx = header_row_idx + df_row_idx + 1

        for df_col, excel_col_idx in col_mapping.items():
            cell = ws.cell(row=excel_row_idx, column=excel_col_idx + 1)
            value = df_row[df_col]

            # Manejar NaN
            if pd.isna(value):
                cell.value = None
            else:
                cell.value = value

def build_monthly_report_path(run_date: date) -> str:
    ensure_all_paths_exist()
    return str(MONTHLY_REPORT_DIR / f"monthly_report_{run_date:%Y-%m-%d}.xlsx")

def apply_to_template(datasets: dict[str, pd.DataFrame], run_date: date | None = None) -> str:
    """
    Copia el template y escribe datasets preservando estilos.
    datasets = {"Hoja": df, ...}
    Devuelve la ruta de salida.
    """
    ensure_all_paths_exist()
    run_date = run_date or date.today()
    template = MONTHLY_REPORT_DIR / "template.xlsx"
    out_path = MONTHLY_REPORT_DIR / f"monthly_report_{run_date:%Y-%m-%d}.xlsx"
    # Copiamos el template para no romper su formato
    shutil.copyfile(template, out_path)
    wb = load_workbook(out_path)
    for sheet, df in datasets.items():
        if sheet not in wb.sheetnames:
            print(f"⚠️ Hoja no encontrada en template: {sheet}")
            continue
        write_df_preserving_style(wb[sheet], df)
        print(f"✅ {sheet}: {len(df)} filas escritas")
    wb.save(out_path)
    print(f"💾 Guardado: {out_path}")
    return str(out_path)
